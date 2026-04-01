from openpyxl import load_workbook
from datetime import datetime, date
import json
import os
import sys

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from core.automation import exec_auto_load

CABECERAS_ESPERADAS = [
    "NOMBRE DEL PROYECTO",
    "TAREA SHORT",
    "TAREA LONG",
    "HORAS",
    "FECHA",
]


def normalizar_fecha(valor):
    if isinstance(valor, datetime):
        return valor.date().isoformat()

    if isinstance(valor, date):
        return valor.isoformat()

    if isinstance(valor, str):
        valor = valor.strip()

        if not valor:
            raise ValueError("Fecha vacía")

        try:
            return datetime.strptime(valor, "%d/%m/%Y").date().isoformat()
        except ValueError:
            pass

        try:
            return datetime.strptime(valor, "%Y-%m-%d").date().isoformat()
        except ValueError:
            pass

        raise ValueError(f"Fecha no válida: {valor}")

    raise ValueError(f"Tipo de fecha no soportado: {type(valor)}")


def fecha_para_crm(fecha_iso):
    dt = datetime.strptime(fecha_iso, "%Y-%m-%d")
    return dt.strftime("%m/%d/%Y")


def limpiar_texto(valor):
    if valor is None:
        return ""
    return str(valor).strip()


def convertir_horas(valor, fila_numero):
    if valor is None or str(valor).strip() == "":
        raise ValueError(f"Fila {fila_numero}: HORAS vacío")

    valor_str = str(valor).strip()

    valor_str = valor_str.replace(",", ".")

    try:
        horas = float(valor_str)
    except ValueError:
        raise ValueError(f"Fila {fila_numero}: HORAS no es un número válido -> {valor}")

    if horas <= 0:
        raise ValueError(f"Fila {fila_numero}: HORAS debe ser mayor que 0")

    if horas.is_integer():
        return int(horas)

    return horas


def validar_cabeceras(ws):
    cabeceras = [
        limpiar_texto(celda)
        for celda in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    ]

    if cabeceras != CABECERAS_ESPERADAS:
        raise ValueError(
            "Las cabeceras del Excel no son correctas.\n"
            f"Esperadas: {CABECERAS_ESPERADAS}\n"
            f"Encontradas: {cabeceras}"
        )


def fila_vacia(fila):
    return all(celda is None or str(celda).strip() == "" for celda in fila)


def resolver_tareas(tarea_short, tarea_long, fila_numero):
    short_limpia = limpiar_texto(tarea_short)
    long_limpia = limpiar_texto(tarea_long)

    if not short_limpia and not long_limpia:
        raise ValueError(f"Fila {fila_numero}: TAREA SHORT y TAREA LONG están vacías")

    if short_limpia and not long_limpia:
        long_limpia = short_limpia

    if long_limpia and not short_limpia:
        short_limpia = long_limpia

    return short_limpia, long_limpia


def leer_registros_excel(ruta_excel):
    wb = load_workbook(ruta_excel, data_only=True)
    ws = wb.active

    validar_cabeceras(ws)

    registros = []

    for fila_numero, fila in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if fila_vacia(fila):
            continue

        if len(fila) < 5:
            raise ValueError(f"Fila {fila_numero}: no tiene las 5 columnas requeridas")

        nombre_proyecto, tarea_short, tarea_long, horas, fecha = fila[:5]

        proyecto_limpio = limpiar_texto(nombre_proyecto)
        if not proyecto_limpio:
            raise ValueError(f"Fila {fila_numero}: NOMBRE DEL PROYECTO vacío")

        tarea_short_final, tarea_long_final = resolver_tareas(
            tarea_short, tarea_long, fila_numero
        )

        horas_num = convertir_horas(horas, fila_numero)
        fecha_iso = normalizar_fecha(fecha)
        fecha_crm = fecha_para_crm(fecha_iso)

        registro = {
            "proyecto": proyecto_limpio,
            "tarea_short": tarea_short_final,
            "tarea_long": tarea_long_final,
            "horas": horas_num,
            "fecha": fecha_crm, # Enviamos la formateada directamente como "fecha"
            "fecha_iso": fecha_iso, # Guardamos la ISO aparte como referencia si hace falta
        }

        registros.append(registro)

    return registros


def main():
    ruta_excel = "suite_excel/registry.xlsx"

    try:
        print(f"--- Leyendo registros de Excel: {ruta_excel} ---")
        registros = leer_registros_excel(ruta_excel)

        exec_auto_load(
            registros=registros,
            headless=False,
            pausa_debug=True,  
        )

    except Exception as e:
        print(f"ERROR: {str(e)}")


if __name__ == "__main__":
    main()